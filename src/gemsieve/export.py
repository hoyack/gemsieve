"""CSV and Excel export for segments, gems, and profiles."""

from __future__ import annotations

import csv
import json
import sqlite3
from pathlib import Path


def export_gems(db: sqlite3.Connection, output_path: str = "gems_export.csv") -> str:
    """Export all gems with explanations to CSV."""
    rows = db.execute(
        """SELECT g.id, g.gem_type, g.sender_domain, g.score, g.explanation,
                  g.recommended_actions, g.status, g.created_at,
                  sp.company_name, sp.industry, sp.company_size
           FROM gems g
           LEFT JOIN sender_profiles sp ON g.sender_domain = sp.sender_domain
           ORDER BY g.score DESC"""
    ).fetchall()

    path = Path(output_path)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "id", "gem_type", "sender_domain", "company_name", "industry",
            "company_size", "score", "summary", "recommended_actions", "status",
        ])
        for row in rows:
            explanation = {}
            try:
                explanation = json.loads(row["explanation"]) if row["explanation"] else {}
            except (json.JSONDecodeError, TypeError):
                pass
            actions = []
            try:
                actions = json.loads(row["recommended_actions"]) if row["recommended_actions"] else []
            except (json.JSONDecodeError, TypeError):
                pass

            writer.writerow([
                row["id"], row["gem_type"], row["sender_domain"],
                row["company_name"] or "", row["industry"] or "",
                row["company_size"] or "", row["score"],
                explanation.get("summary", ""),
                "; ".join(actions), row["status"],
            ])

    return str(path)


def export_segment(
    db: sqlite3.Connection,
    segment: str,
    output_path: str | None = None,
) -> str:
    """Export sender profiles in a given segment to CSV."""
    if output_path is None:
        output_path = f"segment_{segment}.csv"

    rows = db.execute(
        """SELECT sp.*, ss.segment, ss.sub_segment, ss.confidence as segment_confidence
           FROM sender_profiles sp
           JOIN sender_segments ss ON sp.sender_domain = ss.sender_domain
           WHERE ss.segment = ?
           ORDER BY sp.marketing_sophistication_avg""",
        (segment,),
    ).fetchall()

    path = Path(output_path)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "sender_domain", "company_name", "primary_email", "industry",
            "company_size", "marketing_sophistication", "esp_used",
            "product_description", "total_messages", "segment", "sub_segment",
        ])
        for row in rows:
            writer.writerow([
                row["sender_domain"], row["company_name"] or "",
                row["primary_email"] or "", row["industry"] or "",
                row["company_size"] or "", row["marketing_sophistication_avg"] or 0,
                row["esp_used"] or "", row["product_description"] or "",
                row["total_messages"] or 0, row["segment"], row["sub_segment"] or "",
            ])

    return str(path)


def export_all_profiles(
    db: sqlite3.Connection,
    output_path: str = "profiles_export.csv",
    fmt: str = "csv",
) -> str:
    """Export all sender profiles to CSV or Excel."""
    rows = db.execute(
        """SELECT * FROM sender_profiles ORDER BY sender_domain"""
    ).fetchall()

    columns = [
        "sender_domain", "company_name", "primary_email", "reply_to_email",
        "industry", "company_size", "marketing_sophistication_avg",
        "marketing_sophistication_trend", "esp_used", "product_type",
        "product_description", "target_audience", "total_messages",
        "first_contact", "last_contact", "avg_frequency_days",
        "has_personalization", "has_partner_program", "authentication_quality",
    ]

    path = Path(output_path)

    if fmt == "excel":
        if not path.suffix == ".xlsx":
            path = path.with_suffix(".xlsx")
        return _export_excel(rows, columns, str(path))

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        for row in rows:
            writer.writerow([row[col] if col in row.keys() else "" for col in columns])

    return str(path)


def _export_excel(rows: list, columns: list[str], output_path: str) -> str:
    """Export to Excel using openpyxl."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sender Profiles"

    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row[col_name] if col_name in row.keys() else ""
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_path)
    return output_path
